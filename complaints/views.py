from django.db.models import Case, When, IntegerField, Avg
from .models import Complaint, Department, ComplaintFeedback
import csv
import random
from datetime import timedelta
from .utils import create_notification
from django.utils import timezone
from django.db.models import Count, Avg, Case, When, IntegerField
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.paginator import Paginator

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

import openpyxl
from openpyxl.utils import get_column_letter

from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_str
from django.utils.http import urlsafe_base64_decode

from .models import (
    Complaint,
    UserProfile,
    Notification,
    Department,
    DepartmentHead,
    ComplaintCategory,
    ComplaintSubCategory,
    EmailOTP,
    ComplaintFeedback
)
from django.core.mail import EmailMessage
from .forms import ComplaintForm, UserRegistrationForm
from .utils import send_event_email, ADMIN_EMAIL


# =========================
# HOME
# =========================
def home(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        return redirect('dashboard')
    return render(request, 'complaints/home.html')


PUBLIC_ROLE_CHOICES = [

    ('student', 'Student'),

    ('teacher', 'Teacher'),

    ('staff', 'Staff'),

]


@login_required
def analytics_api(request):

    # =========================
    # Complaint Statistics
    # =========================

    total = Complaint.objects.count()

    pending = Complaint.objects.filter(
        status="Pending"
    ).count()

    in_progress = Complaint.objects.filter(
        status="In Progress"
    ).count()

    resolved = Complaint.objects.filter(
        status="Resolved"
    ).count()

    # =========================
    # Quick Statistics
    # =========================

    today = timezone.now()

    today_complaints = Complaint.objects.filter(
        created_at__date=today.date()
    ).count()

    weekly_complaints = Complaint.objects.filter(
        created_at__gte=today - timedelta(days=7)
    ).count()

    monthly_complaints = Complaint.objects.filter(
        created_at__month=today.month,
        created_at__year=today.year
    ).count()

    # =========================
    # Average Resolution Time
    # =========================

    resolved_cases = Complaint.objects.filter(
        status="Resolved",
        resolved_at__isnull=False
    )

    avg_resolution_days = 0

    if resolved_cases.exists():

        total_days = sum(
            (complaint.resolved_at - complaint.created_at).days
            for complaint in resolved_cases
        )

        avg_resolution_days = round(
            total_days / resolved_cases.count(),
            1
        )

    # =========================
    # Notifications
    # =========================

    if request.user.is_superuser:

        notifications = Notification.objects.all()

    else:

        notifications = Notification.objects.filter(
            user=request.user
        )

    notifications = notifications.order_by(
        '-created_at'
    )[:5]
    notification_data = []

    for notification in notifications:

        notification_data.append({
            "title": notification.title,
            "message": notification.message,
            "created_at": notification.created_at.strftime(
                "%d %b %Y %I:%M %p"
            ),
            "is_read": notification.is_read
        })

    # =========================
    # Response
    # =========================

    return JsonResponse({

        "total": total,
        "pending": pending,
        "in_progress": in_progress,
        "resolved": resolved,

        "today_complaints": today_complaints,
        "weekly_complaints": weekly_complaints,
        "monthly_complaints": monthly_complaints,
        "avg_resolution_days": avg_resolution_days,
        "notifications": notification_data

    })
# =========================
# REGISTER
# =========================


def register_view(request):

    if request.method == "POST":

        form = UserRegistrationForm(request.POST)

        if form.is_valid():

            email = form.cleaned_data.get(
                'email'
            )

            mobile = form.cleaned_data.get(
                'mobile_number'
            )

            first_name = form.cleaned_data.get(
                'first_name'
            )

            last_name = form.cleaned_data.get(
                'last_name'
            )

            role = form.cleaned_data.get('role')

            department = form.cleaned_data.get(
                'department'
            )

            if role == "student" and not department:

                form.add_error(
                    'department',
                    'Department is required for students'
                )

                return render(
                    request,
                    "complaints/register.html",
                    {
                        "form": form
                    }
                )
            user = form.save(
                commit=False
            )

            user.is_active = True

            user.email = email

            # Django User table sync
            user.first_name = first_name
            user.last_name = last_name

            user.save()
            UserProfile.objects.create(

                user=user,

                first_name=first_name,

                last_name=last_name,

                email=email,

                department=department,

                mobile_number=mobile,

                role=role

            )

            return render(

                request,

                "complaints/user_created.html"

            )

    else:

        form = UserRegistrationForm()

    return render(

        request,

        "complaints/register.html",

        {

            "form": form

        }

    )

# =========================
# LOGIN / LOGOUT
# =========================


def login_view(request):

    if request.method == 'POST':

        username = request.POST.get(
            'username'
        )

        password = request.POST.get(
            'password'
        )

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user:

            login(request, user)

            profile = UserProfile.objects.filter(
                user=user
            ).first()

            if user.is_superuser:
                return redirect(
                    'admin_dashboard'
                )

            if (
                profile and
                profile.role == "department_head"
            ):
                return redirect(
                    'department_dashboard'
                )

            return redirect(
                'dashboard'
            )

        # Login failed
        return render(
            request,
            'complaints/login.html',
            {
                'error':
                'Invalid username or password'
            }
        )

    return render(
        request,
        'complaints/login.html'
    )


def logout_view(request):
    logout(request)
    return redirect('login')


# =========================
# DASHBOARD (USER)
# =========================
@login_required
def dashboard(request):

    user = request.user

    if user.is_superuser:

        complaints = Complaint.objects.all()

    else:

        profile = UserProfile.objects.filter(
            user=user
        ).first()

        # Department Head
        if profile and profile.role.lower() == "department head":

            complaints = Complaint.objects.filter(
                assigned_to=user
            ).select_related('user', 'assigned_to')

        # Student
        else:

            complaints = Complaint.objects.filter(
                user=user
            )

            dashboard_complaints = complaints.order_by('-id')[:10]

            for c in dashboard_complaints:

                c.show_feedback = False

                if not user.is_superuser:

                    profile = UserProfile.objects.filter(
                        user=user
                    ).first()

                    if profile and profile.role.lower() != "department head":

                        days_open = (
                            timezone.now() - c.created_at
                        ).days

                        if (
                            days_open >= 7
                            and c.status in [
                                "Pending",
                                "In Progress"
                            ]
                        ):
                            c.show_feedback = True

    context = {

        "complaints": dashboard_complaints,

        "total_complaints":
        complaints.count(),

        "pending":
        complaints.filter(
            status="Pending"
        ).count(),

        "in_progress":
        complaints.filter(
            status="In Progress"
        ).count(),

        "resolved":
        complaints.filter(
            status="Resolved"
        ).count()

    }

    return render(
        request,
        "complaints/dashboard.html",
        context
    )


@login_required
def admin_dashboard(request):

    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('dashboard')

    # =========================
    # COMPLAINTS QUERY
    # =========================
    complaints = Complaint.objects.annotate(
        status_order=Case(
            When(status="Pending", then=1),
            When(status="In Progress", then=2),
            When(status="Resolved", then=3),
            default=4,
            output_field=IntegerField()
        )
    ).order_by('status_order', '-id')

    total = complaints.count()
    pending = complaints.filter(status="Pending").count()
    in_progress = complaints.filter(status="In Progress").count()
    resolved = complaints.filter(status="Resolved").count()

    departments = Department.objects.all()
    department_stats = []

    for dept in departments:

        total_dept = Complaint.objects.filter(
            department=dept
        ).count()

        resolved_dept = Complaint.objects.filter(
            department=dept,
            status="Resolved"
        ).count()

        efficiency = 0

        if total_dept > 0:
            efficiency = round(
                (resolved_dept / total_dept) * 100
            )

        department_stats.append({
            "name": dept.name,
            "total": total_dept,
            "resolved": resolved_dept,
            "efficiency": efficiency
        })

    # =========================
    # FEEDBACK DATA
    # =========================
    feedbacks = ComplaintFeedback.objects.all()

    avg_rating = feedbacks.aggregate(avg=Avg('rating'))['avg'] or 0
    avg_satisfaction = feedbacks.aggregate(avg=Avg('satisfaction'))['avg'] or 0

    feedback_count = feedbacks.count()
    # =========================
    # QUICK STATS
    # =========================

    now = timezone.now()
    today = now.date()

    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    today_complaints = Complaint.objects.filter(created_at__date=today).count()
    weekly_complaints = Complaint.objects.filter(
        created_at__date__gte=week_start).count()
    monthly_complaints = Complaint.objects.filter(
        created_at__date__gte=month_start).count()

    notifications = Notification.objects.order_by(
        '-created_at'
    )[:5]

    resolved_complaints = Complaint.objects.filter(
        status='Resolved',
        resolved_at__isnull=False
    )

    avg_resolution_days = 0

    if resolved_complaints.exists():

        total_days = 0

        for c in resolved_complaints:

            if c.created_at and c.resolved_at:

                total_days += (
                    c.resolved_at - c.created_at
                ).days

        avg_resolution_days = round(
            total_days /
            resolved_complaints.count(),
            1
        )

    # =========================
    # FAST FEEDBACK LOOKUP (OPTIMIZED)
    # =========================
    feedback_map = set(
        ComplaintFeedback.objects.values_list('complaint_id', flat=True)
    )

    for c in complaints:
        c.has_feedback_received = c.id in feedback_map
    paginator = Paginator(complaints, 3)

    page_number = request.GET.get('page')

    complaints = paginator.get_page(page_number)
    notifications = Notification.objects.filter(
        user=request.user
    ).order_by(
        '-created_at'
    )[:5]

    # =========================
    # CONTEXT
    # =========================
    return render(request, "complaints/admin_dashboard.html", {
        "total_complaints": total,
        "pending": pending,
        "in_progress": in_progress,
        "resolved": resolved,
        "complaints": complaints,
        "departments": departments,
        "department_stats": department_stats,
        "notifications": notifications,
        "today_complaints": today_complaints,
        "weekly_complaints": weekly_complaints,
        "monthly_complaints": monthly_complaints,
        "avg_resolution_days": avg_resolution_days,

        # ⭐ FEEDBACK METRICS
        "feedbacks": feedbacks,
        "avg_rating": round(avg_rating, 1),
        "avg_satisfaction": round(avg_satisfaction, 1),
        "feedback_count": feedback_count,
    })


@login_required
def department_dashboard(request):

    user = request.user

    # 🔥 ONLY assigned complaints
    complaints = Complaint.objects.filter(
        assigned_to=user
    ).annotate(
        status_order=Case(
            When(status="Pending", then=1),
            When(status="In Progress", then=2),
            When(status="Resolved", then=3),
            default=4,
            output_field=IntegerField()
        )
    ).order_by('status_order', '-id')

    total = complaints.count()
    pending = complaints.filter(status="Pending").count()
    in_progress = complaints.filter(status="In Progress").count()
    resolved = complaints.filter(status="Resolved").count()

    return render(request, "complaints/department_dashboard.html", {
        "complaints": complaints,
        "total_assigned": total,
        "pending": pending,
        "in_progress": in_progress,
        "resolved": resolved
    })
# =========================
# USER DASHBOARD (URL REQUIRED)
# =========================


@login_required
def user_dashboard(request):
    complaints = Complaint.objects.filter(
        user=request.user
    ).order_by('-id')

    return render(request, "complaints/user_dashboard.html", {
        "complaints": complaints,
        "total_complaints": complaints.count()
    })


# =========================
# MY COMPLAINTS
# =========================
@login_required
def my_complaints(request):
    complaints = Complaint.objects.filter(
        user=request.user
    ).order_by('-id')
    return render(request, "complaints/my_complaints.html", {
        "complaints": complaints
    })


# =========================
# REGISTER COMPLAINT
# =========================
@login_required
def register_complaint(request):

    if request.method == "POST":

        form = ComplaintForm(
            request.POST,
            request.FILES,
            user=request.user
        )

        if form.is_valid():

            obj = form.save(commit=False)

            obj.user = request.user

            try:

                profile = UserProfile.objects.get(
                    user=request.user
                )

                obj.first_name = profile.first_name
                obj.last_name = profile.last_name
                obj.email = profile.email
                obj.phone = profile.mobile_number
                obj.role = profile.role

                if not obj.department:
                    obj.department = profile.department

            except UserProfile.DoesNotExist:

                obj.first_name = request.user.first_name
                obj.last_name = request.user.last_name
                obj.email = request.user.email

            obj.save()
            admins = User.objects.filter(is_superuser=True)

            create_notification(
                users=list(admins),
                title="New Complaint",
                message=f"{obj.complaint_id} submitted by {request.user.username}",
                notification_type="system"
            )
            send_event_email(
                event="created",
                complaint=obj,
                recipients=[ADMIN_EMAIL]
            )

            messages.success(
                request,
                "Complaint submitted successfully."
            )

            return redirect("complaint_success")

    else:

        form = ComplaintForm(
            user=request.user
        )

    return render(
        request,
        "complaints/register_complaint.html",
        {
            "form": form,
            "departments": Department.objects.all()
        }
    )
# =========================
# COMPLAINT SUCCESS
# =========================


def complaint_success(request):

    return render(
        request,
        "complaints/success.html"
    )

# =========================
# PROFILE
# =========================


@login_required
def profile_view(request):
    return render(request, "complaints/profile.html")


# =========================
# NOTIFICATIONS
# =========================
@login_required
def mark_notifications_read(request):
    Notification.objects.filter(user=request.user).update(is_read=True)
    return redirect('dashboard')


# =========================
# COMPLAINT DETAIL
# =========================
@login_required
def complaint_detail(request, id):

    complaint = get_object_or_404(
        Complaint,
        id=id
    )

    feedback = ComplaintFeedback.objects.filter(
        complaint=complaint
    ).first()

    return render(
        request,
        "complaints/complaint_detail.html",
        {
            "complaint": complaint,
            "feedback": feedback
        }
    )


# =========================
# ASSIGN COMPLAINT
# =========================
@login_required
def assign_complaint(request, pk):

    complaint = get_object_or_404(
        Complaint,
        pk=pk
    )

    if request.method == "POST":

        print("POST DATA:", request.POST)

        dept_id = request.POST.get(
            "department"
        )

        print(
            "DEPARTMENT ID:",
            dept_id
        )

        if dept_id:

            department = Department.objects.filter(
                id=dept_id
            ).first()

            if department:

                dept_head = DepartmentHead.objects.filter(
                    department=department
                ).first()

                print(
                    "HEAD:",
                    dept_head
                )

                if dept_head:

                    complaint.department = department

                    complaint.assigned_to = dept_head.user

                    complaint.status = "In Progress"

                    complaint.save()

                    admins = User.objects.filter(is_superuser=True)

                    create_notification(
                        users=list(admins),
                        title="Complaint Assigned",
                        message=f"{complaint.complaint_id} assigned to {dept_head.user.username}",
                        notification_type="assignment"
                    )

                    create_notification(
                        users=dept_head.user,
                        title="New Complaint Assigned",
                        message=f"You have been assigned {complaint.complaint_id}",
                        notification_type="assignment"
                    )

                    create_notification(
                        users=complaint.user,
                        title="Complaint Update",
                        message=f"Your complaint {complaint.complaint_id} has been assigned",
                        notification_type="status"
                    )

                    recipients = [complaint.email, ADMIN_EMAIL]

                    if dept_head:
                        recipients.append(dept_head.user.email)

                    send_event_email(
                        event="assigned",
                        complaint=complaint,
                        recipients=recipients
                    )
                    print(
                        "ASSIGNED:",
                        complaint.assigned_to
                    )

        return redirect(
            "admin_dashboard"
        )

    return redirect(
        "admin_dashboard"
    )

# =========================
# UPDATE STATUS
# =========================


@login_required
def update_status(request, pk):
    complaint = get_object_or_404(Complaint, pk=pk)

    status = request.GET.get("status")

    if status in ["Pending", "In Progress", "Resolved"]:
        complaint.status = status
        complaint.save()

        admins = User.objects.filter(is_superuser=True)

        create_notification(
            users=list(admins),
            title="Complaint Status Changed",
            message=f"{complaint.complaint_id} → {status}",
            notification_type="status"
        )

        event_map = {
            "Pending": "status_pending",
            "In Progress": "status_progress",
            "Resolved": "status_resolved"
        }

        event_type = event_map.get(status)

        if event_type:
            send_event_email(
                event=event_type,
                complaint=complaint,
                recipients=[complaint.email, ADMIN_EMAIL]
            )

    return redirect('department_dashboard')


# =========================
# TRACK COMPLAINT (MISSING FIX)
# =========================
def track_complaint(request):
    complaint = None
    error = None
    bar = 0

    if request.method == "POST":
        cid = request.POST.get("complaint_id")

        try:
            complaint = Complaint.objects.get(complaint_id=cid)

            # 🔥 SMART PROGRESS LOGIC
            if complaint.status == "Pending":
                bar = 20
                stage = "pending"

            elif complaint.status == "In Progress":
                bar = 60
                stage = "progress"

            elif complaint.status == "Resolved":
                bar = 100
                stage = "done"

            else:
                bar = 10
                stage = "pending"

        except Complaint.DoesNotExist:
            error = "Complaint ID not found"

    return render(request, "complaints/track_complaint.html", {
        "complaint": complaint,
        "error": error,
        "bar": bar,
    })


# =========================
# CHECK USER (MISSING FIX)
# =========================
def check_user(request):
    email = request.GET.get('email')
    mobile = request.GET.get('mobile')

    return JsonResponse({
        "email_exists": User.objects.filter(email=email).exists(),
        "mobile_exists": UserProfile.objects.filter(mobile_number=mobile).exists()
    })


# =========================
# EMAIL OTP
# =========================

def generate_otp():
    return str(random.randint(100000, 999999))


def send_otp(request):
    email = request.GET.get('email')

    if not email:
        return JsonResponse({
            "status": "error",
            "message": "Email required"
        })

    existing = EmailOTP.objects.filter(
        email=email
    ).order_by('-created_at').first()

    if existing and not existing.can_resend():
        return JsonResponse({
            "status": "error",
            "message": "Please wait before resending OTP"
        })

    otp = generate_otp()

    if existing:
        existing.otp = otp
        existing.attempts = 0
        existing.is_verified = False
        existing.save()

    else:
        EmailOTP.objects.create(
            email=email,
            otp=otp
        )

    subject = "Email Verification - College Complaint Portal"

    message = f"""
    Dear User,

    Welcome to College Complaint Portal.

    To complete your email verification process, please use the OTP below:

    OTP: {otp}

    This OTP will expire in 5 minutes.

    If you did not request this verification, please ignore this email.

    Regards,
    College Complaint Portal Team
    Support Department
    """

    EmailMessage(
        subject,
        message,
        to=[email]
    ).send()

    return JsonResponse({
        "status": "success",
        "message": "OTP sent"
    })


def verify_otp(request):
    email = request.GET.get('email')
    otp = request.GET.get('otp')

    if not email or not otp:
        return JsonResponse({
            "status": "error",
            "message": "Missing data"
        })

    try:

        # latest OTP record
        record = EmailOTP.objects.filter(
            email=email
        ).order_by('-created_at').first()

        if not record:
            return JsonResponse({
                "status": "error",
                "message": "OTP not found"
            })

        if record.is_expired():
            record.delete()

            return JsonResponse({
                "status": "error",
                "message": "OTP expired"
            })

        if record.is_verified:
            return JsonResponse({
                "status": "success",
                "message": "Already verified"
            })

        if record.attempts >= 3:
            record.delete()

            return JsonResponse({
                "status": "error",
                "message": "Too many attempts"
            })

        if str(record.otp).strip() != str(otp).strip():

            record.attempts += 1
            record.save()

            return JsonResponse({
                "status": "error",
                "message": f"Invalid OTP ({3-record.attempts} left)"
            })

        record.is_verified = True
        record.save()

        return JsonResponse({
            "status": "success",
            "message": "OTP Verified"
        })

    except Exception as e:

        return JsonResponse({
            "status": "error",
            "message": str(e)
        })


@login_required
def submit_feedback(request, complaint_id):

    complaint = get_object_or_404(
        Complaint,
        id=complaint_id,
        user=request.user
    )

    if request.method == "POST":

        ComplaintFeedback.objects.update_or_create(
            complaint=complaint,
            defaults={
                "user": request.user,
                "rating": request.POST.get("rating"),
                "resolution_speed": request.POST.get("resolution_speed"),
                "satisfaction": request.POST.get("satisfaction"),
                "comment": request.POST.get("comment"),
            }
        )

        admins = User.objects.filter(is_superuser=True)

        create_notification(
            users=list(admins),
            title="Feedback Received",
            message=f"Feedback submitted for {complaint.complaint_id}",
            notification_type="system"
        )

        if complaint.assigned_to:

            create_notification(
                users=complaint.assigned_to,
                title="Feedback Received",
                message=f"Feedback submitted for {complaint.complaint_id}",
                notification_type="system"
            )

        # Feedback email recipients
        recipients = [ADMIN_EMAIL]

        if complaint.assigned_to and complaint.assigned_to.email:
            recipients.append(
                complaint.assigned_to.email
            )

        # Send feedback notification
        send_event_email(
            event="feedback_submitted",
            complaint=complaint,
            recipients=recipients
        )

        messages.success(
            request,
            "Thank you for your feedback ❤️"
        )

        return redirect("dashboard")

    return render(
        request,
        "complaints/feedback.html",
        {
            "complaint": complaint
        }
    )


@login_required
def complaint_analytics(request):

    if not request.user.is_superuser and not request.user.is_staff:
        return redirect('dashboard')

    complaints = Complaint.objects.all()
    total = complaints.count()
    pending_rate = complaints.filter(status="Pending").count() / max(total, 1)
    progress_rate = complaints.filter(
        status="In Progress").count() / max(total, 1)
    resolved_rate = complaints.filter(
        status="Resolved").count() / max(total, 1)

    old_pending = complaints.filter(
        status="Pending",
        created_at__lt=timezone.now() - timedelta(days=5)
    ).count()

    sla_risk = round((old_pending / max(total, 1)) * 100, 1)

    last_30 = complaints.filter(
        created_at__gte=timezone.now()-timedelta(days=30)).count()
    prev_30 = complaints.filter(
        created_at__gte=timezone.now()-timedelta(days=60),
        created_at__lt=timezone.now()-timedelta(days=30)
    ).count()

    growth = 0 if prev_30 == 0 else round(
        ((last_30 - prev_30) / prev_30) * 100, 1)

    health_score = 100 - (pending_rate * 40) - (sla_risk * 0.5)
    health_score = max(0, round(health_score, 1))

    # 📊 Status wise
    status_data = complaints.values('status').annotate(total=Count('id'))

    # 🏢 Department wise
    department_data = complaints.values(
        'department__name').annotate(total=Count('id'))
    # Department Performance Data
    performance_data = []

    departments = Department.objects.all()

    for dept in departments:

        total_dept = complaints.filter(
            department=dept
        ).count()

        resolved_dept = complaints.filter(
            department=dept,
            status="Resolved"
        ).count()

        performance = 0

        if total_dept > 0:
            performance = round(
                (resolved_dept / total_dept) * 100,
                1
            )

        performance_data.append({
            "name": dept.name,
            "performance": performance
        })

    # 📂 Category wise
    category_data = complaints.values(
        'category__name').annotate(total=Count('id'))

    # 🔢 Summary
    total = complaints.count()
    pending = complaints.filter(status="Pending").count()
    in_progress = complaints.filter(status="In Progress").count()
    resolved = complaints.filter(status="Resolved").count()

    return render(request, "complaints/complaint_analytics.html", {
        "status_data": status_data,
        "department_data": department_data,
        "category_data": category_data,

        "performance_data": performance_data,

        "total_complaints": total,
        "pending": pending,
        "in_progress": in_progress,
        "resolved": resolved,

        "pending_rate": round(pending_rate * 100, 1),
        "progress_rate": round(progress_rate * 100, 1),
        "resolved_rate": round(resolved_rate * 100, 1),

        "sla_risk": sla_risk,
        "growth": growth,
        "health_score": health_score,
    })

# =========================
# EXPORT CSV
# =========================


def export_csv(request):
    complaints = Complaint.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="complaints.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Title', 'Status'])

    for c in complaints:
        writer.writerow([c.id, c.title, c.status])

    return response


# =========================
# EXPORT PDF
# =========================
def export_pdf(request):
    complaints = Complaint.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="complaints.pdf"'

    p = canvas.Canvas(response, pagesize=A4)

    y = 800
    for c in complaints[:30]:
        p.drawString(50, y, f"{c.id} {c.title} {c.status}")
        y -= 20

    p.save()
    return response


def export_excel(request):
    complaints = Complaint.objects.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Complaints"

    headers = ["ID", "Title", "Category", "Status"]
    sheet.append(headers)

    for c in complaints:
        sheet.append([
            c.id,
            c.title,
            str(c.category),
            c.status
        ])

    for col_num, column_title in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=complaints.xlsx'

    workbook.save(response)
    return response


def activate(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except:
        user = None

    if user and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()

        messages.success(
            request, "Email verified successfully! Now you can login.")
        return redirect('login')

    return HttpResponse("Invalid or expired activation link")
# =========================
# AJAX
# =========================


def get_categories(request, dept_id):
    data = list(ComplaintCategory.objects.filter(
        department_id=dept_id).values("id", "name"))
    return JsonResponse(data, safe=False)


def get_subcategories(request, cat_id):
    data = list(ComplaintSubCategory.objects.filter(
        category_id=cat_id).values("id", "name"))
    return JsonResponse(data, safe=False)
